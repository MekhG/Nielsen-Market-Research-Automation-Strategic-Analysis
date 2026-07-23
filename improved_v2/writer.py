# ============================================================
# writer.py
# ============================================================
# FORMAT: Format 2 — Hands-Off Auto-Detection Pipeline
#
# PURPOSE:
# Identical in purpose to Format 1 writer.py — copies the
# output template and populates it with extracted values.
#
# KEY DIFFERENCE FROM FORMAT 1:
#
# FORMAT 1 writer.py:
#   Growth calculation uses period keys from manually
#   configured WIDE_PERIODS / LONG_PERIODS in config.py.
#
# FORMAT 2 writer.py:
#   Growth calculation uses period keys from DetectedConfig
#   object so it works regardless of what periods were found.
#   All other behaviour — dynamic template reading, skip
#   logic, template preservation — is identical to Format 1.
#
# THIS FILE DOES NOT NEED TO BE EDITED BETWEEN RUNS.
# ============================================================

import shutil
import openpyxl
from config import (
    TEMPLATE_FILE, OUTPUT_FILE, TEMPLATE_SHEET,
    TEMPLATE_GEO_ROWS, TEMPLATE_COL_MAP, BLOCK_STARTS,
)


def compute_growth(cy_value, py_value, metric_key):
    """
    Computes growth between current and prior year.
    MS Val% -> percentage point change.
    All others -> percentage growth.
    """
    if cy_value is None or py_value is None:
        return None
    if py_value == 0:
        return None
    if metric_key == "MS Val":
        return round(cy_value - py_value, 2)
    return round(((cy_value - py_value) / py_value) * 100, 2)


def find_geo_rows_dynamic(ws):
    """
    Scans column A to find geography row numbers dynamically.
    Falls back to config coordinates for any not found.
    """
    expected = set(TEMPLATE_GEO_ROWS.keys())
    found    = {}
    for row in ws.iter_rows(min_col=1, max_col=1, values_only=False):
        cell = row[0]
        if cell.value in expected:
            found[cell.value] = cell.row

    missing = expected - set(found.keys())
    if missing:
        print(f"  NOTE: Using config coordinates for: {missing}")
        for geo in missing:
            if geo in TEMPLATE_GEO_ROWS:
                found[geo] = TEMPLATE_GEO_ROWS[geo]

    print(f"  Template scan: {len(found)} geography rows located")
    return found


def _write_sheet(ws, data, geo_rows):
    """Writes all extracted values into the worksheet."""
    written = 0
    skipped = 0

    for (role, geo_label, metric_key, period_key), value in data.items():
        if geo_label not in geo_rows:
            skipped += 1
            continue
        if value is None:
            skipped += 1
            continue
        if (role, metric_key) not in BLOCK_STARTS:
            skipped += 1
            continue
        col = TEMPLATE_COL_MAP.get((role, metric_key, period_key))
        if col is None:
            skipped += 1
            continue

        ws.cell(row=geo_rows[geo_label], column=col).value = round(float(value), 2)
        written += 1

    return written, skipped


def _log_growth(data, geo_rows, detected):
    """
    Logs growth figures using auto-detected period keys.
    Only logs periods that were actually detected in the file.
    """
    print("\n  --- Growth Summary (CY vs PY) — All India ---")
    key_metrics = ["Sales Value", "Sales Units", "MS Val"]
    key_periods = ["Mth", "L3M", "MAT"]
    key_geos    = [g for g in geo_rows if "All India" in str(g)]
    roles       = ["focal", "competitor"]

    for role in roles:
        for geo in key_geos:
            for metric in key_metrics:
                for period in key_periods:
                    cy_key = f"{period}_CY"
                    py_key = f"{period}_PY"
                    # Only log if both periods were detected
                    if cy_key not in detected.periods or py_key not in detected.periods:
                        continue
                    cy     = data.get((role, geo, metric, cy_key))
                    py     = data.get((role, geo, metric, py_key))
                    growth = compute_growth(cy, py, metric)
                    if growth is not None:
                        unit = "pp" if metric == "MS Val" else "%"
                        print(f"    {role:12} | {metric:12} | "
                              f"{period:4} | {growth:+.1f}{unit}")


def write_output(data, detected):
    """
    Copies template, populates it with extracted data,
    logs growth figures using auto-detected period config.

    Args:
        data     : dict from transformer.extract_data()
        detected : DetectedConfig from detector.detect_all()
    """
    print(f"\nWriting output file...")
    print(f"  Template : {TEMPLATE_FILE}")
    print(f"  Output   : {OUTPUT_FILE}")
    print(f"  Sheet    : {TEMPLATE_SHEET}")

    shutil.copy(TEMPLATE_FILE, OUTPUT_FILE)
    wb = openpyxl.load_workbook(OUTPUT_FILE)

    if TEMPLATE_SHEET not in wb.sheetnames:
        raise ValueError(
            f"\n{'='*55}"
            f"\nSHEET NOT FOUND: '{TEMPLATE_SHEET}'"
            f"\nAvailable sheets: {wb.sheetnames}"
            f"\nUpdate TEMPLATE_SHEET in config.py."
            f"\n{'='*55}"
        )

    ws       = wb[TEMPLATE_SHEET]
    geo_rows = find_geo_rows_dynamic(ws)

    written, skipped = _write_sheet(ws, data, geo_rows)
    print(f"  Values written : {written}")
    print(f"  Values skipped : {skipped}")

    _log_growth(data, geo_rows, detected)

    wb.save(OUTPUT_FILE)
    print(f"\n  Output saved: {OUTPUT_FILE}")
