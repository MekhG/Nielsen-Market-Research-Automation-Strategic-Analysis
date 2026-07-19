## Output Writing (writer.py)
## First, dynamic template reading scans the template for
## geography row positions and metric column positions by their labels rather than fixed coordinates,
## making the writer more resilient to minor structural changes in the template. Second, growth
## between current year and prior year is computed and logged for key metrics — percentage point
## change for MS Val% and percentage growth for all other metrics.

%%writefile /content/nielsen/writer.py
import sys
sys.path.append("/content/nielsen")

import shutil
import openpyxl
from config import (
    TEMPLATE_FILE, OUTPUT_FILE,
    TEMPLATE_SHEET_STREPSILS, TEMPLATE_SHEET_DETTOL,
    TEMPLATE_GEO_ROWS, TEMPLATE_COL_MAP, BLOCK_STARTS, PERIOD_OFFSETS
)


# =============================================================================
# GROWTH CALCULATION
# =============================================================================

def compute_growth(cy_value, py_value, metric_key):
    """
    Computes growth between current year and prior year.

    For MS Val — percentage point change (not % growth).
    For all others — percentage growth.

    Returns None if either value is missing or prior year is zero.
    """
    if cy_value is None or py_value is None:
        return None
    if py_value == 0:
        return None

    if metric_key == "MS Val":
        # Percentage point change
        return round(cy_value - py_value, 2)
    else:
        # Percentage growth
        return round(((cy_value - py_value) / py_value) * 100, 2)


# =============================================================================
# DYNAMIC TEMPLATE READING
# =============================================================================

def find_geo_rows(ws):
    """
    Scans column A of the worksheet to find actual row numbers
    for each geography label.
    Works even if rows are inserted or reordered in the template.

    Returns:
        dict of geography label → actual row number
    """
    geo_rows = {}
    expected = set(TEMPLATE_GEO_ROWS.keys())

    for row in ws.iter_rows(min_col=1, max_col=1, values_only=False):
        cell = row[0]
        if cell.value in expected:
            geo_rows[cell.value] = cell.row

    # Warn about any expected geographies not found
    missing = expected - set(geo_rows.keys())
    if missing:
        print(f"  WARNING: These geographies not found in template: {missing}")

    print(f"  Found {len(geo_rows)} geography rows dynamically")
    return geo_rows


def find_metric_cols(ws):
    """
    Scans rows 4 and 5 of the worksheet to find actual column numbers
    for each metric block header.
    Works even if columns are inserted or reordered in the template.

    Returns:
        dict of metric label → actual column number
    """
    expected_metrics = {
        "Sales Value (Cr)", "Sales Units (Mn)", "MS Val %",
        "Store Count (000)", "WD %",
        "Category - Sales Value (Cr)", "Category - Sales Units (Mn)",
        "Category - Store Count (000)",
        "Competitor (GOF) - Sales Value (Cr)",
        "Competitor (GOF) - Sales Units (Mn)",
        "Competitor (GOF) - MS Val %",
        "Competitor (GOF) - Store Count (000)",
        "Competitor (GOF) - WD %"
    }

    metric_cols = {}

    for row in ws.iter_rows(min_row=1, max_row=6, values_only=False):
        for cell in row:
            if cell.value in expected_metrics:
                metric_cols[cell.value] = cell.column

    missing = expected_metrics - set(metric_cols.keys())
    if missing:
        print(f"  WARNING: These metric headers not found in template: {missing}")

    print(f"  Found {len(metric_cols)} metric columns dynamically")
    return metric_cols


# =============================================================================
# WRITER
# =============================================================================

def _write_data_to_sheet(ws, data, use_dynamic=True):
    """
    Writes all extracted values into a single worksheet.

    Args:
        ws          : openpyxl worksheet object
        data        : dict keyed by (role, geo_label, metric_key, period_key) → value
        use_dynamic : if True, scan template dynamically for row/col positions
    """
    # Get geography rows — dynamic or from config
    if use_dynamic:
        geo_rows = find_geo_rows(ws)
    else:
        geo_rows = TEMPLATE_GEO_ROWS

    written = 0
    skipped = 0

    for (role, geo_label, metric_key, period_key), value in data.items():

        # Skip if geography not found
        if geo_label not in geo_rows:
            skipped += 1
            continue

        # Skip if value is None
        if value is None:
            skipped += 1
            continue

        # Skip if role+metric not in template
        if (role, metric_key) not in BLOCK_STARTS:
            skipped += 1
            continue

        # Get row and column
        row = geo_rows[geo_label]
        col = TEMPLATE_COL_MAP.get((role, metric_key, period_key))

        if col is None:
            skipped += 1
            continue

        # Write value
        ws.cell(row=row, column=col).value = round(float(value), 2)
        written += 1

    return written, skipped


def _write_growth_to_sheet(ws, data, geo_rows):
    """
    Computes and prints growth figures for key metrics.
    Growth is not written to template as it has no growth columns,
    but is computed and logged for reference.
    """
    print("\n  --- Growth Summary (CY26 vs PY25) ---")

    key_metrics  = ["Sales Value", "Sales Units", "MS Val"]
    key_periods  = ["Mth", "L3M", "MAT"]
    key_geos     = ["All India (Total)"]
    roles        = ["focal", "competitor"]

    for role in roles:
        for geo in key_geos:
            for metric in key_metrics:
                for period in key_periods:
                    cy = data.get((role, geo, metric, f"{period}_CY"))
                    py = data.get((role, geo, metric, f"{period}_PY"))

                    growth = compute_growth(cy, py, metric)

                    if growth is not None:
                        unit = "pp" if metric == "MS Val" else "%"
                        print(f"    {role:12} | {geo:20} | {metric:12} | "
                              f"{period:4} | {growth:+.1f}{unit}")


def write_output(loz_data, san_data):
    """
    Copies the template and fills it with extracted data.
    Strepsils sheet uses Lozenge data.
    Dettol sheet uses Sanitizer data.
    """
    print("Writing output file...")

    # Copy template — never modify the original
    shutil.copy(TEMPLATE_FILE, OUTPUT_FILE)
    wb = openpyxl.load_workbook(OUTPUT_FILE)

    # --- Strepsils sheet ---
    print("\n  Processing Strepsils sheet...")
    ws_strepsils = wb[TEMPLATE_SHEET_STREPSILS]
    geo_rows_s   = find_geo_rows(ws_strepsils)

    written, skipped = _write_data_to_sheet(ws_strepsils, loz_data)
    print(f"  Strepsils: {written} values written, {skipped} skipped")

    _write_growth_to_sheet(ws_strepsils, loz_data, geo_rows_s)

    # --- Dettol sheet ---
    print("\n  Processing Dettol sheet...")
    ws_dettol  = wb[TEMPLATE_SHEET_DETTOL]
    geo_rows_d = find_geo_rows(ws_dettol)

    written, skipped = _write_data_to_sheet(ws_dettol, san_data)
    print(f"  Dettol: {written} values written, {skipped} skipped")

    _write_growth_to_sheet(ws_dettol, san_data, geo_rows_d)

    # Save
    wb.save(OUTPUT_FILE)
    print(f"\n  Output saved to: {OUTPUT_FILE}")
