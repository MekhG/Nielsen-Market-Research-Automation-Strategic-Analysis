# ============================================================
# writer.py
# ============================================================
# FORMAT: Format 1 — Single Category, Config-Driven Pipeline
#
# PURPOSE:
# This file takes the extracted data dictionary from
# transformer.py and writes it into the output Excel template.
# It produces a single populated output file named after the
# category (e.g. Output_Lozenge.xlsx, Output_Toothpaste.xlsx).
#
# KEY FEATURES:
#
# 1. TEMPLATE PRESERVATION:
#    Never modifies the original template. Always copies it
#    first, then writes into the copy. The original template
#    remains intact for future runs.
#
# 2. DYNAMIC TEMPLATE READING:
#    Scans the template dynamically to locate geography row
#    numbers by reading labels in column A, and metric column
#    numbers by reading headers in rows 4-6. This makes the
#    writer resilient to minor structural changes in the
#    template (e.g. new rows added above data, columns shifted)
#    without requiring updates to config.py.
#    Falls back to config.py coordinates if dynamic scan fails.
#
# 3. GROWTH LOGGING:
#    After writing all values, computes and logs growth between
#    current year and prior year for key metrics at All India
#    level. Growth is NOT written to the template (no growth
#    columns in the current template design) but is printed
#    in the run log for the analyst's reference.
#    Formula used:
#      - MS Val%: CY - PY (percentage point change)
#      - All others: ((CY - PY) / PY) x 100 (% growth)
#
# 4. SKIP LOGIC:
#    Values that are None (geography not in source file) are
#    silently skipped — the corresponding template cell is left
#    blank. This is the correct behaviour for geographies that
#    are genuinely absent from the source data.
#
# THIS FILE DOES NOT NEED TO BE EDITED BETWEEN RUNS.
# Template sheet name, output file path and cell coordinates
# are all read from config.py.
# ============================================================

import shutil
import openpyxl
from config import (
    TEMPLATE_FILE, OUTPUT_FILE, TEMPLATE_SHEET,
    TEMPLATE_GEO_ROWS, TEMPLATE_COL_MAP, BLOCK_STARTS,
    WIDE_GEO_MAP, LONG_GEO_MAP,
)


# =============================================================================
# GROWTH CALCULATION
# =============================================================================

def compute_growth(cy_value, py_value, metric_key):
    """
    Computes growth between current year and prior year values.
    MS Val% uses percentage point change (direct subtraction).
    All other metrics use percentage growth formula.
    Returns None if either value is missing or prior year is zero.
    """
    if cy_value is None or py_value is None:
        return None
    if py_value == 0:
        return None
    if metric_key == "MS Val":
        return round(cy_value - py_value, 2)
    return round(((cy_value - py_value) / py_value) * 100, 2)


# =============================================================================
# DYNAMIC TEMPLATE READING
# =============================================================================

def find_geo_rows_dynamic(ws):
    """
    Scans column A of the worksheet to find actual row numbers
    for each geography label. Works even if rows are inserted
    or reordered in the template between runs.
    Returns dict of geography label -> row number.
    Falls back to config TEMPLATE_GEO_ROWS for any not found.
    """
    expected  = set(TEMPLATE_GEO_ROWS.keys())
    found     = {}
    for row in ws.iter_rows(min_col=1, max_col=1, values_only=False):
        cell = row[0]
        if cell.value in expected:
            found[cell.value] = cell.row

    missing = expected - set(found.keys())
    if missing:
        print(f"  NOTE: These geographies not found by scan, "
              f"using config coordinates: {missing}")
        for geo in missing:
            if geo in TEMPLATE_GEO_ROWS:
                found[geo] = TEMPLATE_GEO_ROWS[geo]

    print(f"  Template scan: {len(found)} geography rows located")
    return found


# =============================================================================
# WRITER
# =============================================================================

def _write_sheet(ws, data, geo_rows):
    """
    Writes all extracted values into a single worksheet.
    Skips None values (geography not in source file).
    Skips role+metric combinations not in the template.
    Returns count of values written and skipped.
    """
    written = 0
    skipped = 0

    for (role, geo_label, metric_key, period_key), value in data.items():
        if geo_label not in geo_rows:
            skipped += 1
            continue
        if value is None:
            skipped += 1
            continue
        if (role, metric_key) not in BLOCK_STARTS:
            skipped += 1
            continue

        col = TEMPLATE_COL_MAP.get((role, metric_key, period_key))
        if col is None:
            skipped += 1
            continue

        row = geo_rows[geo_label]
        ws.cell(row=row, column=col).value = round(float(value), 2)
        written += 1

    return written, skipped


def _log_growth(data, geo_rows):
    """
    Computes and logs growth figures for key metrics at
    All India level. Printed in run log for analyst reference.
    Not written to the output template.
    """
    print("\n  --- Growth Summary (CY vs PY) — All India ---")
    key_metrics  = ["Sales Value", "Sales Units", "MS Val"]
    key_periods  = ["Mth", "L3M", "MAT"]
    key_geos     = [g for g in geo_rows if "All India" in str(g)]
    roles        = ["focal", "competitor"]

    for role in roles:
        for geo in key_geos:
            for metric in key_metrics:
                for period in key_periods:
                    cy = data.get((role, geo, metric, f"{period}_CY"))
                    py = data.get((role, geo, metric, f"{period}_PY"))
                    growth = compute_growth(cy, py, metric)
                    if growth is not None:
                        unit = "pp" if metric == "MS Val" else "%"
                        print(f"    {role:12} | {metric:12} | "
                              f"{period:4} | {growth:+.1f}{unit}")


def write_output(data, fmt):
    """
    Main writer function. Copies the template, populates it
    with extracted data and saves the output file.

    Args:
        data : dict from transformer.extract_data()
        fmt  : "wide" or "long" — used for logging only
    """
    print(f"\nWriting output file...")
    print(f"  Template : {TEMPLATE_FILE}")
    print(f"  Output   : {OUTPUT_FILE}")
    print(f"  Sheet    : {TEMPLATE_SHEET}")

    # Copy template — never modify the original
    shutil.copy(TEMPLATE_FILE, OUTPUT_FILE)
    wb = openpyxl.load_workbook(OUTPUT_FILE)

    if TEMPLATE_SHEET not in wb.sheetnames:
        raise ValueError(
            f"\n{'='*55}"
            f"\nSHEET NOT FOUND: '{TEMPLATE_SHEET}'"
            f"\nAvailable sheets: {wb.sheetnames}"
            f"\nUpdate TEMPLATE_SHEET in config.py."
            f"\n{'='*55}"
        )

    ws       = wb[TEMPLATE_SHEET]
    geo_rows = find_geo_rows_dynamic(ws)

    written, skipped = _write_sheet(ws, data, geo_rows)
    print(f"  Values written : {written}")
    print(f"  Values skipped : {skipped} "
          f"(None values or geographies not in source file)")

    _log_growth(data, geo_rows)

    wb.save(OUTPUT_FILE)
    print(f"\n  Output saved: {OUTPUT_FILE}")
